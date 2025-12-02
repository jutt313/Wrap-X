"""Utility helpers to extract readable text previews from uploaded documents."""
from __future__ import annotations

import base64
import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)


TEXT_EXTENSIONS = {
    "txt",
    "md",
    "markdown",
    "csv",
    "json",
    "xml",
    "log",
    "yaml",
    "yml",
    "ini",
}

TEXT_MIME_PREFIXES = ("text/",)
TEXT_MIME_EXTRAS = {
    "application/json",
    "application/xml",
}


def _decode_base64(data: str) -> Optional[bytes]:
    try:
        return base64.b64decode(data)
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Failed to decode base64 document: %s", exc)
        return None


def _is_text_like(file_type: Optional[str], mime_type: Optional[str]) -> bool:
    ft = (file_type or "").lower()
    mime = (mime_type or "").lower()
    if ft in TEXT_EXTENSIONS:
        return True
    if any(mime.startswith(prefix) for prefix in TEXT_MIME_PREFIXES):
        return True
    if mime in TEXT_MIME_EXTRAS:
        return True
    return False


def _extract_text_bytes(raw: bytes, encoding: str = "utf-8") -> Optional[str]:
    try:
        return raw.decode(encoding, errors="ignore")
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Failed to decode text document: %s", exc)
        return None


def _extract_pdf(raw: bytes) -> Optional[str]:
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        pages = []
        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text.strip())
        return "\n".join(filter(None, pages)) or None
    except ImportError:
        logger.warning("PyPDF2 not installed; skipping PDF extraction")
    except Exception as exc:  # pragma: no cover - best-effort
        logger.warning("PDF extraction failed: %s", exc)
    return None


def _extract_docx(raw: bytes) -> Optional[str]:
    try:
        import docx  # type: ignore

        document = docx.Document(io.BytesIO(raw))
        paragraphs = [para.text.strip() for para in document.paragraphs if para.text]
        return "\n".join(paragraphs) or None
    except ImportError:
        logger.warning("python-docx not installed; skipping DOCX extraction")
    except Exception as exc:
        logger.warning("DOCX extraction failed: %s", exc)
    return None


def _extract_xlsx(raw: bytes, max_rows: int = 50, max_cols: int = 10) -> Optional[str]:
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        sheet = wb.active
        rows = []
        for ridx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if ridx > max_rows:
                break
            values = []
            for cidx, cell in enumerate(row, start=1):
                if cidx > max_cols:
                    break
                if cell is None:
                    continue
                values.append(str(cell))
            if values:
                rows.append(", ".join(values))
        wb.close()
        return "\n".join(rows) or None
    except ImportError:
        logger.warning("openpyxl not installed; skipping XLSX extraction")
    except Exception as exc:
        logger.warning("XLSX extraction failed: %s", exc)
    return None


def _clean_preview(text: str, max_chars: int) -> str:
    collapsed = " ".join(text.split())
    if len(collapsed) > max_chars:
        return collapsed[:max_chars].rstrip() + "…"
    return collapsed


def extract_text_preview(
    *,
    content_b64: Optional[str],
    file_type: Optional[str],
    mime_type: Optional[str],
    max_chars: int = 2000,
) -> Optional[str]:
    """Attempt to extract a readable preview (2000 chars default) from a document."""
    if not content_b64:
        return None

    raw = _decode_base64(content_b64)
    if not raw:
        return None

    text: Optional[str] = None
    ft = (file_type or "").lower()

    if _is_text_like(file_type, mime_type):
        text = _extract_text_bytes(raw)
    elif ft in {"pdf"}:
        text = _extract_pdf(raw)
    elif ft in {"docx"}:
        text = _extract_docx(raw)
    elif ft in {"xlsx", "xlsm"}:
        text = _extract_xlsx(raw)
    else:
        # fall back to naive utf-8 decode; may still succeed for many formats
        text = _extract_text_bytes(raw)

    if not text:
        return None
    cleaned = _clean_preview(text, max_chars)
    return cleaned or None


def extract_full_text(
    *,
    content_b64: Optional[str],
    file_type: Optional[str],
    mime_type: Optional[str],
) -> Optional[str]:
    """Extract the full text content from a document (no truncation)."""
    if not content_b64:
        return None

    raw = _decode_base64(content_b64)
    if not raw:
        return None

    text: Optional[str] = None
    ft = (file_type or "").lower()

    if _is_text_like(file_type, mime_type):
        text = _extract_text_bytes(raw)
    elif ft in {"pdf"}:
        text = _extract_pdf(raw)
    elif ft in {"docx"}:
        text = _extract_docx(raw)
    elif ft in {"xlsx", "xlsm"}:
        # For Excel, extract more rows/cols for full content
        text = _extract_xlsx_full(raw)
    else:
        # fall back to naive utf-8 decode; may still succeed for many formats
        text = _extract_text_bytes(raw)

    if not text:
        return None
    
    # Return full text, cleaned but not truncated
    return text.strip() or None


def _extract_xlsx_full(raw: bytes, max_rows: int = 10000, max_cols: int = 100) -> Optional[str]:
    """Extract full content from Excel file (larger limits)."""
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        sheet = wb.active
        rows = []
        for ridx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if ridx > max_rows:
                rows.append(f"... (truncated at {max_rows} rows)")
                break
            values = []
            for cidx, cell in enumerate(row, start=1):
                if cidx > max_cols:
                    values.append("...")
                    break
                if cell is None:
                    continue
                values.append(str(cell))
            if values:
                rows.append(", ".join(values))
        wb.close()
        return "\n".join(rows) or None
    except ImportError:
        logger.warning("openpyxl not installed; skipping XLSX extraction")
    except Exception as exc:
        logger.warning("XLSX extraction failed: %s", exc)
    return None
